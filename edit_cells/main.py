#OMMA Lic data 2021
import openpyxl

# Abre el archivo Excel existente
archivo_excel = "ex.xlsx"
workbook = openpyxl.load_workbook(archivo_excel)

# Selecciona la hoja de trabajo en la que quieres trabajar
hoja = workbook.active

# Definir las filas de inicio y fin (ten en cuenta que Python usa índices basados en cero)
fila_inicio = 318  # La fila 319 en base a índices en Python
fila_fin = 2345   # La fila 2346 en base a índices en Python

# Recorre las filas seleccionadas
for fila in range(fila_inicio, fila_fin + 1):
    # Obtiene el contenido de la celda en la primera columna
    contenido_celda = hoja.cell(row=fila + 1, column=1).value
    
    # Divide el contenido en dos partes utilizando "Trade Name:"
    partes = contenido_celda.split("Trade Name:")
    
    # Actualiza la primera columna con la primera parte (sin "Trade Name:")
    hoja.cell(row=fila + 1, column=1, value=partes[0].strip())
    
    # Actualiza la tercera columna con la segunda parte (sin "Trade Name:")
    hoja.cell(row=fila + 1, column=3, value=partes[1].strip() if len(partes) > 1 else "")

# Guarda los cambios en un nuevo archivo
workbook.save("copia_de_tu_archivo.xlsx")

# Cierra el archivo Excel
workbook.close()


