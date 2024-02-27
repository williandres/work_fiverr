import pdfplumber
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import fitz  # PyMuPDF
import os

carpeta = "/home/willian/Repositories/work_fiverr/pdf_tables/James Pdfs/Pdfs"
wb = Workbook()
archivos = os.listdir(carpeta)
pdfs = [archivo for archivo in archivos if archivo.endswith(".pdf")]
rutas_pdfs = [os.path.join(carpeta, pdf) for pdf in pdfs]
rutas_pdfs = sorted(rutas_pdfs)

for file_path in rutas_pdfs:
    #Extraccion
    with pdfplumber.open(file_path) as pdf:
        hipervinculos = []
        tabla = []
        for i in pdf.pages:
            page = i.extract_table()
            if page is not None:
                for j in page:
                    tabla.append(j)

    doc = fitz.open(file_path)
    for i, page in enumerate(doc):
        links = page.get_links()
        if links:
            for link in links:
                hipervinculos.append(link)
    doc.close()

    #Modificacion
    #Eliminar '/n
    for fila in tabla:
        for i, elemento in enumerate(fila):
            fila[i] = elemento.replace('\n', ' ')#
    #Enumerar
    tabla[0].insert(0, 'Number')
    for i, fila in enumerate(tabla[1:], start=1):
        fila.insert(0, str(i))

    #Crear hoja Excel
    ws = wb.create_sheet(title=file_path.split('/')[-1])
    for row in tabla:
        ws.append(row)
    #Hipervinculos
    n = 2
    for i, elemento in enumerate(hipervinculos):
        ws[f'F{i+n}'] = "Click"
        ws[f'F{i+n}'].hyperlink = elemento['uri']
        ws[f'F{i+n}'].style = "Hyperlink"
        ws[f'F{i+n}'].font = Font(underline="single", color="0563C1")

wb.save("james_pdfs.xlsx")